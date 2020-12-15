import datetime
import MySQLdb
import mysql.connector
from mysql.connector import Error
from openpyxl import load_workbook

#try:

connection = MySQLdb.connect(
     host="localhost",
     db="volkszaehler",
     user="vz-admin", passwd="secure"
    )
#sql_select_Query = "select * from data"
cursor = connection.cursor()
#cursor.execute(sql_select_Query)
#records = cursor.fetchall()
#print("Total number of rows in data is: ", cursor.rowcount)
sql_select_Query = "SELECT * FROM data ORDER BY ID DESC LIMIT 1"
cursor.execute(sql_select_Query)
records = cursor.fetchall()
for row in records:
    print(row[3])
    local_zaehlerstand=row[3]

mySql_insert_query = """INSERT INTO daily (zaehlerstand, info) VALUES (%s, %s) """
recordTuple = (float(local_zaehlerstand),"")
cursor.execute(mySql_insert_query, recordTuple)
connection.commit()

sql_select_Query = "SELECT (SELECT zaehlerstand FROM daily ORDER BY ID DESC Limit 0,1) - (SELECT zaehlerstand FROM daily ORDER BY ID DESC Limit 1,1) AS Differenz"
cursor.execute(sql_select_Query)
records = cursor.fetchall()
for row in records:
    print(row)
    local_differenz=row[0]

cursor.close()

datetime_object = datetime.date.today()

wb2= load_workbook("/mnt/NAS/share/Stromverbrauch.xlsx")
print(wb2.sheetnames)
ws=wb2[str(datetime_object.month)+"-"+str(datetime_object.year)]
ws.append([str(datetime_object.day)+"."+str(datetime_object.month)+"."+str(datetime_object.year),local_zaehlerstand,local_differenz])
#ws.append([1,local_zaehlerstand,local_differenz])

wb2.save("/mnt/NAS/share/Stromverbrauch.xlsx")
